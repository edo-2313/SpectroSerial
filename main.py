from openpyxl import Workbook   #modulo per manipolare file excel
from openpyxl.chart import (    #moduli per creare grafici su excel
    LineChart,
    Reference,
)
import os
import serial
import re
altro='S'

re1='(start)'	# Word 1
re2='(\\s+)'	# White Space 1
re3='(time)'	# Word 2

end = re.compile(re1+re2+re3,re.IGNORECASE|re.DOTALL)
stringa=[]
print('~=~=Spettrofotometro=~=~')
print('')
porta = serial.Serial('COM4')     #apri porta seriale
wb = Workbook()                 #crea un file excel
wb.remove(wb.active)            #rimuove il foglio di default
while altro=='S':
    nome = input('Nome del campione: ')         #inserire il nome del campione da analizzare
    ws = wb.create_sheet(nome)
    dati = []                                   #crea l'input
    print('\nAttendo i dati...')
    while True:
        dati = porta.read()                #legge 1 byte
        if dati==bytes('\r','utf-8'):
            stringa = ''.join(dati)             #converte la lista in stringa
            if stringa=='WL    AU':
                scan=True                       #la scansione è iniziata!
                print('iniziato!')
            elif end.search(stringa):
                scan=False                     #trovato la stringa con "start time..." quindi la scansione è finita
                print('finito')     
            else:
                a,b=stringa.split(": ")
            print(a+'---'+b)
            stringa=[]
        elif dati==bytes('\n','utf-8'):
            pass
        else:
            stringa.append(str(dati, 'utf-8'))
    for a in range(1,10):
        for b in range(1,10):
            ws.cell(row=a, column=b, value=a+b)
    grafico = LineChart()
    grafico.title = "Curva assorbanza di "+nome
    grafico.style = 13
    grafico.y_axis.title = 'Assorbanza'
    grafico.x_axis.title = 'Lunghezza d\'onda'
    altro = 'C'
    while altro!='S' and altro!='N':
        altro = input('Inserire un altro campione? (S/N)')

wb.save(os.environ['UserProfile']+'\\Desktop\\dati.xlsx')
