from openpyxl import Workbook   #modulo per manipolare file excel
from openpyxl.chart import (    #moduli per creare grafici su excel
    ScatterChart,
    Reference,
    Series,
)
import os
import re
import datetime
altro='S'
now=datetime.datetime.now()
re1='(start)'	# Word 1
re2='(\\s+)'	# White Space 1
re3='(time)'	# Word 2

end = re.compile(re1+re2+re3,re.IGNORECASE|re.DOTALL)
print('~=~=Spettrofotometro=~=~')
print('')
try:
    file = open("log.txt","r")     #apri file
except:
    print('File non trovato!')
    time.sleep(5)
wb = Workbook()                 #crea un file excel
wb.remove(wb.active)            #rimuove il foglio di default
while altro=='S':
    nome = input('Nome del campione: ')         #inserire il nome del campione da analizzare
    ws = wb.create_sheet(nome)
    ws.cell(row=1,column=1,value="Lunghezza")
    ws.cell(row=1,column=2,value="Assorbanza")
    print('\nAttendo i dati...')
    linea=[]
    scan=False
    riga=2
    while True:
        byte = bytes(file.read(1),'utf-8')                #legge 1 byte
        #print(byte)
        if byte==b"\n":
            stringa = ''.join(linea)             #converte la lista in stringa
            #print('rip')
            if end.search(stringa):
                scan=False                     #trovato la stringa con "start time..." quindi la scansione è finita
                print('Scansione terminata!')
                break
            if scan==True:
                stringa = stringa.replace(" ","")
                a,b=stringa.split(":")
                #print(a+'---'+b)
                ws.cell(row=riga,column=1,value=int(a))
                ws.cell(row=riga,column=2,value=float(b))
                riga=riga+1
            if stringa=='WL    AU':
                scan=True                       #la scansione è iniziata!
                print('Inizio scansione!')
            linea=[]
        elif byte!=b"\r":
            linea.append(str(byte, 'utf-8'))
    assorbanze=[]
    for i in range(2,riga):
        assorbanze.append(ws["B"+str(i)].value)
    lunghezze=[]
    for i in range(2,riga):
        lunghezze.append(ws["A"+str(i)].value)
    grafico = ScatterChart()
    grafico.title = "Curva assorbanza di "+nome
    grafico.style = 13
    grafico.y_axis.title = 'Assorbanza'
    grafico.x_axis.title = 'Lunghezza d\'onda'
    grafico.legend=None
    grafico.x_axis.scaling.min=min(lunghezze)
    grafico.x_axis.scaling.max=max(lunghezze)
    grafico.y_axis.scaling.min=min(assorbanze)
    grafico.y_axis.scaling.max=max(assorbanze)
    valorix=Reference(ws, min_col=1,min_row=2,max_row=riga-1)
    valoriy=Reference(ws, min_col=2,min_row=2,max_row=riga-1)
    serie=Series(valoriy,valorix,title_from_data=True)
    grafico.series.append(serie)
    ws.add_chart(grafico)    
    altro = 'C'
    while altro!='S' and altro!='N':
        altro = input('Scansionare un altro campione? (S/N)')
        altro = altro.replace("s","S")
        altro = altro.replace("n","N")
nomefile = input('\nNome del file: ')
print('SALVATAGGIO... IL FILE SARÀ SALVATO SUL DESKTOP!')
wb.save(os.environ['UserProfile']+'\\Desktop\\'+nomefile+'.xlsx')
print('SALVATO')
